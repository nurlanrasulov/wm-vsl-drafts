#!/usr/bin/env python3
"""Generate shrink report from Snowflake and print top 10 preview."""

from __future__ import annotations

import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

import send_shrink_report as report
from shrink_report.week_utils import last_week_start


def main() -> None:
    report.load_dotenv()
    data_source = report.resolve_data_source("auto")
    week_start = last_week_start()
    output_dir = ROOT / "output" / "shrink-reports"
    name, data = report.generate_report(
        output_dir=output_dir,
        week_start=week_start,
        data_source=data_source,
    )

    sheet = load_workbook(io.BytesIO(data)).active
    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    headers = rows[0] if rows else []
    print(f"\nGenerated: {output_dir / name}")
    print("Columns:", ", ".join(str(h) for h in headers if h))
    print("\nTop contributors:")
    for row in rows[1:]:
        print(" | ".join("" if value is None else str(value) for value in row))


if __name__ == "__main__":
    main()
