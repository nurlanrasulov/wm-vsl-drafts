"""Excel post-processing for vendor reports."""

from __future__ import annotations

import io

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def format_vsl_by_item_report(data: bytes, column_name: str) -> bytes:
    """Bold the VSL column and sort rows by it descending."""
    workbook = load_workbook(io.BytesIO(data))
    sheet = _find_sheet_with_column(workbook, column_name)
    if sheet is None:
        raise RuntimeError(f'Column "{column_name}" not found in Excel export')

    header_row = 1
    col_idx = _find_column_index(sheet, header_row, column_name)
    if col_idx is None:
        raise RuntimeError(f'Column "{column_name}" not found in sheet "{sheet.title}"')

    bold_font = Font(bold=True)
    for row_idx in range(header_row, sheet.max_row + 1):
        sheet.cell(row=row_idx, column=col_idx).font = bold_font

    _sort_sheet_by_column_desc(sheet, col_idx, header_row=header_row)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _find_sheet_with_column(workbook, column_name: str):
    target = column_name.strip().lower()
    for sheet in workbook.worksheets:
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col).value
            if value and str(value).strip().lower() == target:
                return sheet
    return None


def _find_column_index(sheet, header_row: int, column_name: str) -> int | None:
    target = column_name.strip().lower()
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=header_row, column=col).value
        if value and str(value).strip().lower() == target:
            return col
    return None


def _sort_sheet_by_column_desc(sheet, col_idx: int, *, header_row: int) -> None:
    max_row = sheet.max_row
    max_col = sheet.max_column
    if max_row <= header_row:
        return

    rows: list[tuple[tuple, list]] = []
    for row_idx in range(header_row + 1, max_row + 1):
        row_values = [sheet.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        sort_key = _sort_key(row_values[col_idx - 1])
        rows.append((sort_key, row_values))

    rows.sort(key=lambda item: item[0], reverse=True)

    for offset, (_, row_values) in enumerate(rows, start=header_row + 1):
        for col in range(1, max_col + 1):
            sheet.cell(row=offset, column=col).value = row_values[col - 1]

    # Preserve column widths after sort (no-op if unset).
    _ = get_column_letter(col_idx)


def _sort_key(value) -> tuple:
    if value is None:
        return (1, "")
    if isinstance(value, (int, float)):
        return (0, float(value))
    text = str(value).strip().replace("%", "")
    try:
        return (0, float(text))
    except ValueError:
        return (0, text.lower())
