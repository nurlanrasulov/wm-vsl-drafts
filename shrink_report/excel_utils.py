"""Excel post-processing for shrink contributor reports."""

from __future__ import annotations

import io

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def format_shrink_report(
    data: bytes,
    *,
    contributor_column: str,
    category_column: str,
    gtin_column: str,
    excluded_categories: tuple[str, ...],
    top_n: int = 10,
    output_column_order: tuple[str, ...] | None = None,
) -> bytes:
    """
    Exclude Herbs and FnV, sort by shrink contribution descending (biggest contributors first),
    keep only the top N rows, and reorder columns with GTIN included.
    """
    workbook = load_workbook(io.BytesIO(data))
    sheet = _find_sheet_with_column(workbook, contributor_column)
    if sheet is None:
        raise RuntimeError(f'Column "{contributor_column}" not found in Excel export')

    header_row = 1
    contributor_col = _find_column_index(sheet, header_row, contributor_column)
    if contributor_col is None:
        raise RuntimeError(
            f'Column "{contributor_column}" not found in sheet "{sheet.title}"'
        )

    gtin_col = _find_column_index(sheet, header_row, gtin_column)
    if gtin_col is None:
        raise RuntimeError(
            f'Column "{gtin_column}" not found in Excel export. '
            "Ensure the Looker explore includes GTIN or set SHRINK_REPORT_GTIN_COLUMN."
        )

    category_col = _find_column_index(sheet, header_row, category_column)

    max_col = sheet.max_column
    max_row = sheet.max_row
    if max_row <= header_row:
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    headers = [sheet.cell(row=header_row, column=col).value for col in range(1, max_col + 1)]
    excluded = {value.strip().lower() for value in excluded_categories}
    rows: list[tuple[tuple, list]] = []

    for row_idx in range(header_row + 1, max_row + 1):
        row_values = [
            sheet.cell(row=row_idx, column=col).value for col in range(1, max_col + 1)
        ]
        if category_col is not None:
            category_value = row_values[category_col - 1]
            if category_value is not None and str(category_value).strip().lower() in excluded:
                continue

        sort_key = _sort_key(row_values[contributor_col - 1])
        rows.append((sort_key, row_values))

    rows.sort(key=lambda item: item[0], reverse=True)

    if top_n > 0:
        rows = rows[:top_n]

    if output_column_order:
        headers, row_data = _reorder_columns(headers, [row for _, row in rows], output_column_order)
    else:
        row_data = [row for _, row in rows]

    bold_font = Font(bold=True)
    contributor_header_idx = _header_index(headers, contributor_column)
    gtin_header_idx = _header_index(headers, gtin_column)

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = bold_font

    for row_offset, row_values in enumerate(row_data, start=header_row + 1):
        for col_idx, value in enumerate(row_values, start=1):
            sheet.cell(row=row_offset, column=col_idx).value = value

    for row_idx in range(header_row + 1 + len(row_data), max_row + 1):
        for col in range(1, max_col + 1):
            sheet.cell(row=row_idx, column=col).value = None

    for trailing_col in range(len(headers) + 1, max_col + 1):
        for row_idx in range(header_row, max_row + 1):
            sheet.cell(row=row_idx, column=trailing_col).value = None

    if contributor_header_idx is not None:
        _ = get_column_letter(contributor_header_idx + 1)
    if gtin_header_idx is not None:
        _ = get_column_letter(gtin_header_idx + 1)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _reorder_columns(
    headers: list,
    rows: list[list],
    column_order: tuple[str, ...],
) -> tuple[list, list[list]]:
    normalized_headers = {
        str(header).strip().lower(): idx for idx, header in enumerate(headers) if header
    }
    ordered_indices: list[int] = []

    for column_name in column_order:
        idx = normalized_headers.get(column_name.strip().lower())
        if idx is not None and idx not in ordered_indices:
            ordered_indices.append(idx)

    for idx in range(len(headers)):
        if idx not in ordered_indices:
            ordered_indices.append(idx)

    new_headers = [headers[idx] for idx in ordered_indices]
    new_rows = [[row[idx] for idx in ordered_indices] for row in rows]
    return new_headers, new_rows


def _header_index(headers: list, column_name: str) -> int | None:
    target = column_name.strip().lower()
    for idx, header in enumerate(headers):
        if header and str(header).strip().lower() == target:
            return idx
    return None


def _find_sheet_with_column(workbook, column_name: str):
    target = column_name.strip().lower()
    for sheet in workbook.worksheets:
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col).value
            if value and str(value).strip().lower() == target:
                return sheet
    return None


def _find_column_index(sheet, header_row: int, column_name: str) -> int | None:
    target = column_name.strip().lower()
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=header_row, column=col).value
        if value and str(value).strip().lower() == target:
            return col
    return None


def _sort_key(value) -> tuple:
    if value is None:
        return (1, float("-inf"))
    if isinstance(value, (int, float)):
        return (0, float(value))
    text = str(value).strip().replace("%", "").replace(",", "")
    try:
        return (0, float(text))
    except ValueError:
        return (0, float("-inf"))
